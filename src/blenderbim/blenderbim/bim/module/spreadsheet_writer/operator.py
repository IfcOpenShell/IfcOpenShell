import os
import sys
import time
import site
import collections
import subprocess

site.addsitedir(os.path.join(os.path.dirname(os.path.realpath(__file__)), "libs", "site", "packages"))

import bpy
from bpy.props import StringProperty, BoolProperty, IntProperty, EnumProperty
from bpy_extras.io_utils import ImportHelper 
from bpy.types import (Operator, PropertyGroup)

import blenderbim.bim.import_ifc
from blenderbim.bim.ifc import IfcStore
import blenderbim.tool as tool
import ifcopenshell

import openpyxl
from openpyxl import load_workbook
import pandas as pd
import xlsxwriter
import zipfile
import xml.parsers.expat

from collections import defaultdict
from collections import OrderedDict


#print ('openpyxl', openpyxl.__version__, openpyxl.__file__)
#print ('pandas',pd.__version__, pd.__file__)
#print ('xlsxwriter',xlsxwriter.__version__, xlsxwriter.__file__)

#https://stackoverflow.com/questions/72657415/fix-futurewarning-related-to-the-pandas-append-function

#function to open file on windows, mac and linux
def open_file(filename):
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.call([opener, filename])
 
 
        
class Element(list):
    def __init__(self, name, attrs):
        self.name = name
        self.attrs = attrs


class TreeBuilder:
    def __init__(self):
        self.root = Element("root", None)
        self.path = [self.root]
    def start_element(self, name, attrs):
        element = Element(name, attrs)
        self.path[-1].append(element)
        self.path.append(element)
    def end_element(self, name):
        assert name == self.path[-1].name
        self.path.pop()
    def char_data(self, data):
        self.path[-1].append(data)


def get_hidden_rows(node):
    row = 0
    for e in node:
        if not isinstance(e, Element):
            continue
        yield from get_hidden_rows(e)
        if e.name != "table:table-row":
            continue
        attrs = e.attrs
        rows = int(attrs.get("table:number-rows-repeated", 1))
        if "table:visibility" in attrs.keys():  # If the key is here, we can assume it's hidden (or can we ?)
            #for row_idx in range(rows):
            #    yield row + row_idx
            yield from range(row, row + rows)
        row += rows

        
        

class ConstructDataFrame:

    def __init__(self, context):
        
        blenderbim_spreadsheet_properties = context.scene.blenderbim_spreadsheet_properties
        my_collection = context.scene.my_collection
        
        ifc_dictionary = defaultdict(list)
        ifc_custom_property_list = []
        
        blenderbim_spreadsheet_properties.my_workbook = 'Overview'
        blenderbim_spreadsheet_properties.my_xlsx_file  = IfcStore.path.replace('.ifc','_blenderbim.xlsx') 
        blenderbim_spreadsheet_properties.my_ods_file = IfcStore.path.replace('.ifc','_blenderbim.ods') 

        
        ifc_file = ifcopenshell.open(IfcStore.path)
        products = ifc_file.by_type('IfcProduct')
        
        ifc_version = (ifc_file.schema)
        print ("IFC version found: ", ifc_version)
        
        is_external = 'IsExternal'
        loadbearing = 'LoadBearing'
        fire_rating = 'FireRating'
        
        length = 'Length'
        width = 'Width'
        height = 'Height'
        area = 'Area' 
        netarea = 'NetArea'
        netsidearea = 'NetSideArea'
        volume = 'Volume'
        netvolume = 'NetVolume'
        perimeter = 'Perimeter'

        for product in products:
            ##################################################################
            ##########################  GlobalId #############################
            ##################################################################
            ifc_dictionary['GlobalId'].append(str(product.GlobalId))
             
            ##################################################################
            ########################### General ##############################
            ##################################################################
            if blenderbim_spreadsheet_properties.my_ifcproduct:
                ifc_dictionary['IfcProduct'].append(str(product.is_a()))
                
            if blenderbim_spreadsheet_properties.my_ifcbuildingstorey:  
                ifc_dictionary['IfcBuildingStorey'].append(self.get_ifcbuildingstorey(context, ifcproduct=product, ifc_version=ifc_version)[0])    
    
            if blenderbim_spreadsheet_properties.my_ifcproduct_name: 
                ifc_dictionary['Name'].append(str(product.Name))
                
            if blenderbim_spreadsheet_properties.my_type:     
                ifc_dictionary['Type'].append(self.get_ifcproducttype_name(context, ifcproduct=product)[0])
            
            if blenderbim_spreadsheet_properties.my_ifcclassification: 
                ifc_dictionary['Classification'].append(self.get_classification_code(context, ifcproduct=product, ifc_version=ifc_version)[0])
              
            if blenderbim_spreadsheet_properties.my_ifcmaterial:     
                ifc_dictionary['Material(s)'].append(self.get_materials(context, ifcproduct=product)[0])
                
            
            ##################################################################
            ################### Common Properties ############################
            ##################################################################  
            if blenderbim_spreadsheet_properties.my_isexternal:    
                ifc_dictionary[is_external].append(self.get_common_properties(context, ifcproduct=product, property_name=is_external)[0])
                
            if blenderbim_spreadsheet_properties.my_loadbearing:     
                ifc_dictionary[loadbearing].append(self.get_common_properties(context, ifcproduct=product, property_name=loadbearing)[0])
                 
            if blenderbim_spreadsheet_properties.my_firerating:    
                ifc_dictionary[fire_rating].append(self.get_common_properties(context, ifcproduct=product, property_name=fire_rating)[0])
               
            ##################################################################
            ##################### Base Quantities ############################
            ##################################################################
            if blenderbim_spreadsheet_properties.my_length:  
                ifc_dictionary[length].append(self.get_quantities(context, ifcproduct=product, quantity_name=length)[0])

            if blenderbim_spreadsheet_properties.my_width:  
                ifc_dictionary[width].append(self.get_quantities(context, ifcproduct=product, quantity_name=width)[0])   
                
            if blenderbim_spreadsheet_properties.my_height: 
                ifc_dictionary[height].append(self.get_quantities(context, ifcproduct=product, quantity_name=height)[0])   
                  
            if blenderbim_spreadsheet_properties.my_area:     
                ifc_dictionary[area].append(self.get_quantities(context, ifcproduct=product, quantity_name=area)[0])
                
            if blenderbim_spreadsheet_properties.my_netarea:     
                ifc_dictionary[netarea].append(self.get_quantities(context, ifcproduct=product, quantity_name=netarea)[0])
                
            if blenderbim_spreadsheet_properties.my_netsidearea:     
                ifc_dictionary[netsidearea].append(self.get_quantities(context, ifcproduct=product, quantity_name=netsidearea)[0])
                
            if blenderbim_spreadsheet_properties.my_volume:   
                ifc_dictionary[volume].append(self.get_quantities(context, ifcproduct=product, quantity_name=volume)[0]) 
                
            if blenderbim_spreadsheet_properties.my_netvolume:   
                ifc_dictionary[netvolume].append(self.get_quantities(context, ifcproduct=product, quantity_name=netvolume)[0])
                
            if blenderbim_spreadsheet_properties.my_perimeter:    
                ifc_dictionary[perimeter].append(self.get_quantities(context, ifcproduct=product, quantity_name=perimeter)[0])
   
            ##################################################################
            ################### Custom Properties ############################
            ##################################################################
            if len(context.scene.my_collection.items) > 1:
                for item in context.scene.my_collection.items:
                    ifc_dictionary[item.name].append(self.get_custom_pset( context,ifcproduct=product,
                                                                    pset_name=str(item.name).split('.')[0],
                                                                    property_name=str(item.name).split('.')[1])[0])   
      
        df = pd.DataFrame(ifc_dictionary)
        self.df = df
        
        
    def get_ifcproducttype_name(self, context, ifcproduct):
        
        type_name_list = []
        
        # NOTE: The product representations are defined as representation maps
        # (at the level of the supertype IfcTypeProduct, which gets assigned by an element occurrence
        # instance through the IfcShapeRepresentation.Item[1] being an IfcMappedItem.
        
        if ifcproduct:
            
            ifcproduct_type = ifcopenshell.util.element.get_type(ifcproduct)
            
            if ifcproduct_type:
                #ifcproduct_type.Name
                type_name_list.append(ifcproduct_type.Name)
            
                #only applies to labels
                #type_name_list.append(ifcproduct.ObjectType)
            
        if not type_name_list:
            type_name_list.append(None)
    
        
        return type_name_list
    
    def get_ifcbuildingstorey(self, context, ifcproduct, ifc_version):
        building_storey_list = []
        
        try:
            if ifcproduct:  
               
                #maybe for future reference
                if ifc_version == 'IFC2X3':
                    if ifcproduct.ContainedInStructure:
                        for contained_in_structure in (ifcproduct.ContainedInStructure):
                            building_storey_list.append(contained_in_structure.RelatingStructure.Name)
                                
                                
                #maybe for future reference               
                if ifc_version == 'IFC4':
                    if ifcproduct.ContainedInStructure:
                        for contained_in_structure in (ifcproduct.ContainedInStructure):
                            building_storey_list.append(contained_in_structure.RelatingStructure.Name)
                            
                            
        except:
            pass
        #IfcOpeningElement should not be linked directly to the spatial structure of the project,
        #i.e. the inverse relationship ContainedInStructure shall be NIL.
        #It is assigned to the spatial structure through the elements it penetrates.
        if not building_storey_list:
            building_storey_list.append(None)
             
        return building_storey_list 
    
    def get_classification_code(self, context, ifcproduct, ifc_version):
    
        #Classifications of an object may be referenced from an external source rather than being
        #contained within the IFC model. This is done through the IfcClassificationReference class.
        #ClassificationRefForObjects', 
        
        assembly_code_list = []

        if ifcproduct.HasAssociations:
     
            if ifcproduct.HasAssociations[0].is_a() == 'IfcRelAssociatesClassification':
              
                if ifcproduct.HasAssociations[0].RelatingClassification:
                    
                    #for IFC2x3
                    if ifc_version == 'IFC2X3':
                        if ifcproduct.HasAssociations[0].RelatingClassification.ItemReference:
                            assembly_code_list.append(ifcproduct.HasAssociations[0].RelatingClassification.ItemReference)
                       
                    #for IFC4 
                    if ifc_version == 'IFC4':
                        if ifcproduct.HasAssociations[0].RelatingClassification.Identification:
                            assembly_code_list.append(ifcproduct.HasAssociations[0].RelatingClassification.Identification)
                            
            if ifcproduct.HasAssociations[0].is_a() == 'IfcRelAssociatesMaterial':
            
                for i in ifcproduct.HasAssociations:
                    if i.is_a() == 'IfcRelAssociatesClassification':
                        
                        #for IFC2x3
                        if ifc_version == 'IFC2X3':
                            if i.RelatingClassification.ItemReference:
                                assembly_code_list.append(i.RelatingClassification.ItemReference)
                            
                        #for IFC4     
                        if ifc_version == 'IFC4':
                            if i.RelatingClassification:
                                assembly_code_list.append(i.RelatingClassification.Identification)
                
                                               
        if not assembly_code_list:
            assembly_code_list.append(None)
       
        return assembly_code_list
    
    def get_materials(self, context, ifcproduct):
        
        material_list = []
        
        if ifcproduct.HasAssociations:
            for i in ifcproduct.HasAssociations:
                if i.is_a('IfcRelAssociatesMaterial'):
                    
                    if i.RelatingMaterial.is_a('IfcMaterial'):
                        material_list.append(i.RelatingMaterial.Name)
                      
                    #IfcMaterialList deprecated since IFC 4? 
                    if i.RelatingMaterial.is_a('IfcMaterialList'):
                        for materials in i.RelatingMaterial.Materials:
                            material_list.append(materials.Name)
                             
                    if i.RelatingMaterial.is_a('IfcMaterialLayerSetUsage'):
                        for materials in i.RelatingMaterial.ForLayerSet.MaterialLayers:
                            material_list.append(materials.Material.Name)
                            
                    else:
                        pass
                          
        if not material_list:
            material_list.append('None')
           
        joined_material_list = ' | '.join(material_list)
                             
        return [joined_material_list] 
    
    def get_common_properties(self, context, ifcproduct, property_name):
        
        pset_common_list = []
     
        if ifcproduct.IsDefinedBy:        
            for ifcreldefinesbyproperties in ifcproduct.IsDefinedBy:
                if (ifcreldefinesbyproperties.is_a()) == 'IfcRelDefinesByProperties':
                    if ifcreldefinesbyproperties.RelatingPropertyDefinition.is_a() == 'IfcPropertySet':
                        if (ifcreldefinesbyproperties.RelatingPropertyDefinition.Name).endswith('Common'):
                            
                            for ifcproperty in (ifcreldefinesbyproperties.RelatingPropertyDefinition.HasProperties):
                                if (ifcproperty.Name == property_name):
                                    if ifcproperty.NominalValue:
                                    
                                        pset_common_list.append(ifcproperty.NominalValue[0])
                               
                                        
        if not pset_common_list:
            pset_common_list.append(None)  

        return pset_common_list
      
    def get_quantities(self, context, ifcproduct, quantity_name):
        
        quantity_list = []
        length_value_list = ['Length','Width','Height','Perimeter']
        
        for properties in ifcproduct.IsDefinedBy:
            if properties.is_a('IfcRelDefinesByProperties'):
                if properties.RelatingPropertyDefinition.is_a('IfcElementQuantity'):
                    for quantities in properties.RelatingPropertyDefinition.Quantities:
                        if (quantities.Name) == quantity_name:
                            
                            if quantities.Name in length_value_list:
                                quantity_list.append(float(quantities.LengthValue))
                                
                            if quantity_name == 'Area':
                                quantity_list.append(float(quantities.AreaValue))
                                
                            if quantity_name == 'NetArea':
                                quantity_list.append(float(quantities.AreaValue))
                            
                            if quantity_name == 'NetSideArea':
                                quantity_list.append(float(quantities.AreaValue))
                                
                            if quantity_name == 'Volume':
                                quantity_list.append(float(quantities.VolumeValue))
                                
                            if quantity_name == 'NetVolume':
                                quantity_list.append(float(quantities.VolumeValue))
                                                
        if not quantity_list:
            quantity_list.append(None)
                  
        return quantity_list
    
    def get_custom_pset(self, context, ifcproduct, pset_name, property_name):
        
        custom_pset_list = []
        
        if ifcproduct.IsDefinedBy:        
            for ifcreldefinesbyproperties in ifcproduct.IsDefinedBy:
                if (ifcreldefinesbyproperties.is_a()) == 'IfcRelDefinesByProperties':
                    if ifcreldefinesbyproperties.RelatingPropertyDefinition.is_a() == 'IfcPropertySet':
                        if pset_name in (ifcreldefinesbyproperties.RelatingPropertyDefinition.Name):
                            for ifcproperty in (ifcreldefinesbyproperties.RelatingPropertyDefinition.HasProperties):
                                
                                if (ifcproperty.Name == property_name):
                                    custom_pset_list.append(ifcproperty.NominalValue[0])
                      
                                        
        if not custom_pset_list:
            custom_pset_list.append(None)
                             
        return custom_pset_list
    
    
    
class WriteToXLSX(bpy.types.Operator):
    """Write IFC data to .xlsx"""
    bl_idname = "object.write_to_xlsx"
    bl_label = "Simple Object Operator"
    

    def execute(self, context):
        print("Write to .xlsx")
        start_time = time.perf_counter()
        
        construct_data_frame = ConstructDataFrame(context)
  
        blenderbim_spreadsheet_properties = context.scene.blenderbim_spreadsheet_properties
        blenderbim_spreadsheet_properties.my_workbook
        print ('xlsx file: ', blenderbim_spreadsheet_properties.my_xlsx_file)
        
        writer = pd.ExcelWriter(blenderbim_spreadsheet_properties.my_xlsx_file, engine='xlsxwriter')
        construct_data_frame.df.to_excel(writer, sheet_name=blenderbim_spreadsheet_properties.my_workbook, startrow=1, header=False, index=False)
        
        workbook  = writer.book
        #cell_format = workbook.add_format({'bold': True,'border': 1,'bg_color': '#4F81BD','font_color': 'white','font_size':14})
        
        worksheet = writer.sheets[blenderbim_spreadsheet_properties.my_workbook]
        #worksheet.write('A1', str(IfcStore.path), cell_format)
        
        (max_row, max_col) = construct_data_frame.df.shape
         
        # Create a list of column headers, to use in add_table().
        column_settings = []
        for header in construct_data_frame.df.columns:
            column_settings.append({'header': header})

        # Add the table.
        worksheet.add_table(1, 0, max_row, max_col - 1, {'columns': column_settings})

        # Make the columns wider for clarity.
        worksheet.set_column(0, max_col - 1, 30)
          
        #find out from the pandas dataframe in which column the calculation needs to be positioned.
        for header_name in construct_data_frame.df.columns:
            if (header_name == 'Area'):
                col_no = construct_data_frame.df.columns.get_loc("Area")
                column_letter = (xlsxwriter.utility.xl_col_to_name(col_no))
                
                #Works in MS Excel, but not in LibreOffice
                #="Area: " &SUBTOTAL(109;F2:F3821)
               
                total_area='=SUBTOTAL(109,' + str(column_letter) + '3:' + str(column_letter) + str(construct_data_frame.df[construct_data_frame.df.columns[0]].count()) + ')'
                worksheet.write_formula(str(column_letter)+'1', total_area)
                
            if header_name == 'NetArea':
                col_no = construct_data_frame.df.columns.get_loc("NetArea")
                column_letter = (xlsxwriter.utility.xl_col_to_name(col_no))
                total_area='=SUBTOTAL(109,' + str(column_letter) + '3:' + str(column_letter) + str(construct_data_frame.df[construct_data_frame.df.columns[0]].count()) + ')'
                worksheet.write_formula(str(column_letter)+'1', total_area)    
                
            if header_name == 'NetSideArea':
                col_no = construct_data_frame.df.columns.get_loc("NetSideArea")
                column_letter = (xlsxwriter.utility.xl_col_to_name(col_no))
                total_area='=SUBTOTAL(109,' + str(column_letter) + '3:' + str(column_letter) + str(construct_data_frame.df[construct_data_frame.df.columns[0]].count()) + ')'
                worksheet.write_formula(str(column_letter)+'1', total_area)
                  
            if header_name == 'Volume':
                col_no = construct_data_frame.df.columns.get_loc("Volume")
                column_letter = (xlsxwriter.utility.xl_col_to_name(col_no))
                
                total_volume='=SUBTOTAL(109,' + str(column_letter) + '3:' + str(column_letter) + str(construct_data_frame.df[construct_data_frame.df.columns[0]].count()) + ')'
                worksheet.write_formula(str(column_letter)+'1', total_volume)
                
            if header_name == 'NetVolume':
                col_no = construct_data_frame.df.columns.get_loc("NetVolume")
                column_letter = (xlsxwriter.utility.xl_col_to_name(col_no))
                
                total_volume='=SUBTOTAL(109,' + str(column_letter) + '3:' + str(column_letter) + str(construct_data_frame.df[construct_data_frame.df.columns[0]].count()) + ')'
                worksheet.write_formula(str(column_letter)+'1', total_volume)
                
        # Close the Pandas Excel writer and output the Excel file.
        writer.save()
        open_file(blenderbim_spreadsheet_properties.my_xlsx_file)
        
        print (time.perf_counter() - start_time, "seconds for the .xlsx to be written")
        
        blenderbim_spreadsheet_properties.my_file_path = IfcStore.path.replace('.ifc','_blenderbim.xlsx')
   
        return {'FINISHED'}
    
    

    
class WriteToODS(bpy.types.Operator):
    """Write IFC data to .ods"""
    bl_idname = "object.write_to_ods"
    bl_label = "Simple ODS Object Operator"
    
    
    def execute(self, context):
        
        print("Write to .ods")
        
        start_time = time.perf_counter()
        construct_data_frame = ConstructDataFrame(context)
  
        blenderbim_spreadsheet_properties = context.scene.blenderbim_spreadsheet_properties
        blenderbim_spreadsheet_properties.my_workbook
        print ('ods file: ', blenderbim_spreadsheet_properties.my_ods_file)
        
        writer_ods = pd.ExcelWriter(blenderbim_spreadsheet_properties.my_ods_file, engine='odf')
        construct_data_frame.df.to_excel(writer_ods, sheet_name=blenderbim_spreadsheet_properties.my_workbook, startrow=0, header=True, index=False)
        
        worksheet_ods = writer_ods.sheets[blenderbim_spreadsheet_properties.my_workbook]
        writer_ods.save()
   
        open_file(blenderbim_spreadsheet_properties.my_ods_file)
        blenderbim_spreadsheet_properties.my_file_path = IfcStore.path.replace('.ifc','_blenderbim.ods')
        
        print (time.perf_counter() - start_time, "seconds for the .ods to be written")

        
        return {'FINISHED'}

       
class FilterIFCElements(bpy.types.Operator):
    """Show the IFC elements you filtered in the spreadsheet"""
    bl_idname = "object.filter_ifc_elements"
    bl_label = "select the IFC elements"
    

    def execute(self, context):
   
        blenderbim_spreadsheet_properties = context.scene.blenderbim_spreadsheet_properties 
        
        if blenderbim_spreadsheet_properties.my_file_path == "":
            print ("Use the 'Write IFC data to .xlsx' or 'Write IFC data to .ods' buttons first.")
        
        if blenderbim_spreadsheet_properties.my_file_path:
            if blenderbim_spreadsheet_properties.my_file_path.endswith(".xlsx"):
                print ("Retrieving data from: " , blenderbim_spreadsheet_properties.my_file_path)
            
                blenderbim_spreadsheet_properties.my_workbook = 'Overview'
                workbook_openpyxl = load_workbook(blenderbim_spreadsheet_properties.my_file_path)
                worksheet_openpyxl = workbook_openpyxl[blenderbim_spreadsheet_properties.my_workbook] 
                
                global_id_filtered_list = []
                       
                for row_idx in range(3, worksheet_openpyxl.max_row + 1):
                    if not worksheet_openpyxl.row_dimensions[row_idx].hidden:
                        cell = worksheet_openpyxl[f"A{row_idx}"]
                        global_id_filtered_list.append(str(cell.value))

                self.filter_IFC_elements(context, guid_list=global_id_filtered_list)
                
                return {'FINISHED'}

                
            if blenderbim_spreadsheet_properties.my_file_path.endswith(".ods"):
                print ("Retrieving data from: " , blenderbim_spreadsheet_properties.my_file_path)
            
                # Get content xml data from OpenDocument file
                ziparchive = zipfile.ZipFile(blenderbim_spreadsheet_properties.my_file_path, "r")
                xmldata = ziparchive.read("content.xml")
                ziparchive.close()
                
                # Create parser and parsehandler
                parser = xml.parsers.expat.ParserCreate()
                treebuilder = TreeBuilder()
                # Assign the handler functions
                parser.StartElementHandler  = treebuilder.start_element
                parser.EndElementHandler    = treebuilder.end_element
                parser.CharacterDataHandler = treebuilder.char_data

                # Parse the data
                parser.Parse(xmldata, True)

                hidden_rows = get_hidden_rows(treebuilder.root)  # This returns a generator object
        
                dataframe = pd.read_excel(blenderbim_spreadsheet_properties.my_file_path, sheet_name=blenderbim_spreadsheet_properties.my_workbook, skiprows=hidden_rows, engine="odf")
                self.filter_IFC_elements(context, guid_list=dataframe['GlobalId'].values.tolist())
                
                return {'FINISHED'}
                
                
    def filter_IFC_elements(self, context, guid_list):
        
        print ("Filtering IFC elements")
        
        start_time = time.perf_counter()
        
        outliner = next(a for a in bpy.context.screen.areas if a.type == "OUTLINER") 
        outliner.spaces[0].show_restrict_column_viewport = not outliner.spaces[0].show_restrict_column_viewport
        
        bpy.ops.object.select_all(action='DESELECT')
      
        for obj in bpy.context.view_layer.objects:
            element = tool.Ifc.get_entity(obj)
            if element is None:        
                obj.hide_viewport = True
                continue
            data = element.get_info()
       
            
            obj.hide_viewport = data.get("GlobalId", False) not in guid_list

        bpy.ops.object.select_all(action='SELECT') 
        
        print (time.perf_counter() - start_time, "seconds to show the IFC elements")
            

        
        
    
class UnhideIFCElements(bpy.types.Operator):
    """Unhide all IFC elements"""
    bl_idname = "object.unhide_all"
    bl_label = "Unhide All"

    def execute(self, context):
        print("Unhide all")
        
        for obj in bpy.data.objects:
            obj.hide_viewport = False 
        
        return {'FINISHED'}  
 
 
class MyCollectionActions(bpy.types.Operator):
    bl_idname = "my.collection_actions"
    bl_label = "Execute"
    action: bpy.props.EnumProperty(
        items=(
            ("add",) * 3,
            ("remove",) * 3,
        ),
    )
    def execute(self, context):
        my_collection = context.scene.my_collection
        if self.action == "add":           
            item = my_collection.items.add()  
        if self.action == "remove":
            my_collection.items.remove(len(my_collection.items) - 1)
        return {"FINISHED"}