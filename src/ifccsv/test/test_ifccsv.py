# This file was generated with the assistance of an AI coding tool.
#
# IfcCSV - A utility to interact with IFC data through CSV.
# Copyright (C) 2020, 2021 Dion Moult <dion@thinkmoult.com>
#
# This file is part of IfcCSV.
#
# IfcCSV is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# IfcCSV is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with IfcCSV.  If not, see <http://www.gnu.org/licenses/>.

import csv
from pathlib import Path

import ifcopenshell
import ifcopenshell.api.root
import openpyxl
import pandas as pd
import pytest
from odf.opendocument import load as load_ods
from odf.table import Table

import ifccsv


def build_ifc_file() -> ifcopenshell.file:
    ifc_file = ifcopenshell.file()
    ifcopenshell.api.root.create_entity(ifc_file, ifc_class="IfcProject")
    ifcopenshell.api.root.create_entity(ifc_file, ifc_class="IfcWall", name="Wall1")
    ifcopenshell.api.root.create_entity(ifc_file, ifc_class="IfcWall", name="Wall2")
    ifcopenshell.api.root.create_entity(ifc_file, ifc_class="IfcSlab", name="Slab1")
    ifcopenshell.api.root.create_entity(ifc_file, ifc_class="IfcDoor", name="Door1")
    return ifc_file


class TestSplitBy:
    def get_elements(self, ifc_file: ifcopenshell.file) -> list[ifcopenshell.entity_instance]:
        return ifc_file.by_type("IfcElement")

    def test_split_results_partitions_rows_by_attribute_value(self):
        ifc_file = build_ifc_file()
        elements = self.get_elements(ifc_file)

        ifc_csv = ifccsv.IfcCsv()
        ifc_csv.export(ifc_file, elements, ["class", "Name"], format="pd", split_by="class")

        assert ifc_csv.split_groups is not None
        split_keys = {key for key, _ in ifc_csv.split_groups}
        assert split_keys == {"IfcWall", "IfcSlab", "IfcDoor"}

        rows_by_class = dict(ifc_csv.split_groups)
        assert len(rows_by_class["IfcWall"]) == 2
        assert len(rows_by_class["IfcSlab"]) == 1
        assert len(rows_by_class["IfcDoor"]) == 1

    def test_split_by_unknown_attribute_raises(self):
        ifc_file = build_ifc_file()
        elements = self.get_elements(ifc_file)

        ifc_csv = ifccsv.IfcCsv()
        with pytest.raises(ValueError):
            ifc_csv.export(ifc_file, elements, ["class", "Name"], format="pd", split_by="NotAnAttribute")

    def test_export_pd_without_split_returns_single_dataframe(self):
        ifc_file = build_ifc_file()
        elements = self.get_elements(ifc_file)

        ifc_csv = ifccsv.IfcCsv()
        df = ifc_csv.export(ifc_file, elements, ["class", "Name"], format="pd")

        assert isinstance(df, pd.DataFrame)
        assert len(df) == 4

    def test_export_pd_with_split_returns_dict_of_dataframes(self):
        ifc_file = build_ifc_file()
        elements = self.get_elements(ifc_file)

        ifc_csv = ifccsv.IfcCsv()
        result = ifc_csv.export(ifc_file, elements, ["class", "Name"], format="pd", split_by="class")

        assert isinstance(result, dict)
        assert set(result.keys()) == {"IfcWall", "IfcSlab", "IfcDoor"}
        assert len(result["IfcWall"]) == 2
        assert list(result["IfcWall"].columns) == ["GlobalId", "class", "Name"]

    def test_export_xlsx_writes_one_worksheet_per_split_value(self, tmp_path: Path):
        ifc_file = build_ifc_file()
        elements = self.get_elements(ifc_file)
        output = tmp_path / "data.xlsx"

        ifc_csv = ifccsv.IfcCsv()
        ifc_csv.export(ifc_file, elements, ["class", "Name"], output=str(output), format="xlsx", split_by="class")

        workbook = openpyxl.load_workbook(output)
        assert set(workbook.sheetnames) == {"IfcWall", "IfcSlab", "IfcDoor"}

        wall_sheet = workbook["IfcWall"]
        rows = list(wall_sheet.iter_rows(values_only=True))
        assert rows[0] == ("GlobalId", "class", "Name")
        assert len(rows) == 3  # header + 2 walls

    def test_export_ods_writes_one_worksheet_per_split_value(self, tmp_path: Path):
        ifc_file = build_ifc_file()
        elements = self.get_elements(ifc_file)
        output = tmp_path / "data.ods"

        ifc_csv = ifccsv.IfcCsv()
        ifc_csv.export(ifc_file, elements, ["class", "Name"], output=str(output), format="ods", split_by="class")

        document = load_ods(str(output))
        tables = document.spreadsheet.getElementsByType(Table)
        sheet_names = {table.getAttribute("name") for table in tables}
        assert sheet_names == {"IfcWall", "IfcSlab", "IfcDoor"}

    def test_export_csv_writes_one_suffixed_file_per_split_value(self, tmp_path: Path):
        ifc_file = build_ifc_file()
        elements = self.get_elements(ifc_file)
        output = tmp_path / "data.csv"

        ifc_csv = ifccsv.IfcCsv()
        ifc_csv.export(ifc_file, elements, ["class", "Name"], output=str(output), format="csv", split_by="class")

        assert not output.exists()
        wall_csv = tmp_path / "data_IfcWall.csv"
        slab_csv = tmp_path / "data_IfcSlab.csv"
        door_csv = tmp_path / "data_IfcDoor.csv"
        assert wall_csv.exists()
        assert slab_csv.exists()
        assert door_csv.exists()

        with wall_csv.open(newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert rows[0] == ["GlobalId", "class", "Name"]
        assert len(rows) == 3  # header + 2 walls

    def test_export_csv_without_split_writes_single_file(self, tmp_path: Path):
        ifc_file = build_ifc_file()
        elements = self.get_elements(ifc_file)
        output = tmp_path / "data.csv"

        ifc_csv = ifccsv.IfcCsv()
        ifc_csv.export(ifc_file, elements, ["class", "Name"], output=str(output), format="csv")

        assert output.exists()
        with output.open(newline="", encoding="utf-8") as f:
            rows = list(csv.reader(f))
        assert len(rows) == 5  # header + 4 elements


class TestSanitizeSplitKey:
    def test_strips_forbidden_worksheet_characters(self):
        existing: set[str] = set()
        name = ifccsv.IfcCsv.sanitize_split_key("Pset/Foo:Bar*[1]", existing)
        assert all(char not in name for char in "\\/*?:[]")

    def test_truncates_to_max_length(self):
        existing: set[str] = set()
        long_value = "x" * 50
        name = ifccsv.IfcCsv.sanitize_split_key(long_value, existing, max_length=31)
        assert len(name) <= 31

    def test_disambiguates_collisions(self):
        existing: set[str] = set()
        first = ifccsv.IfcCsv.sanitize_split_key("Wall", existing)
        second = ifccsv.IfcCsv.sanitize_split_key("Wall", existing)
        assert first != second

    def test_empty_value_falls_back_to_placeholder(self):
        existing: set[str] = set()
        name = ifccsv.IfcCsv.sanitize_split_key("", existing)
        assert name == "Sheet"
