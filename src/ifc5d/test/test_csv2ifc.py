# IfcOpenShell - IFC toolkit and geometry engine
# Copyright (C) 2022 Dion Moult <dion@thinkmoult.com>
#
# This file is part of IfcOpenShell.
#
# IfcOpenShell is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# IfcOpenShell is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with IfcOpenShell.  If not, see <http://www.gnu.org/licenses/>.

import csv
import json
import tempfile
from pathlib import Path

import ifcopenshell
import ifcopenshell.api.root
import ifcopenshell.api.unit
import ifcopenshell.util.cost
import pytest

import ifc5d.csv2ifc
import ifc5d.ifc5Dspreadsheet


class TestCsv2Ifc:
    @staticmethod
    def setup_ifc_file() -> ifcopenshell.file:
        ifc_file = ifcopenshell.file()
        ifcopenshell.api.root.create_entity(ifc_file, ifc_class="IfcProject")
        unit = ifcopenshell.api.unit.add_si_unit(ifc_file, unit_type="LENGTHUNIT", prefix="MILLI")
        ifcopenshell.api.unit.assign_unit(ifc_file, units=[unit])
        return ifc_file

    @staticmethod
    def get_items_rows_from_csv(csv_filepath: Path, delimiter: str = ",") -> list[list[str]]:
        with csv_filepath.open("r", encoding="utf-8") as csv_file:
            reader = csv.reader(csv_file, delimiter=delimiter)
            rows = list(reader)
        rows = rows[1:]  # Skip header.
        rows = filter(lambda l: l[0], rows)  # Skip empty rows.
        rows = list(rows)
        return rows

    @staticmethod
    def validate_ifc_file_against_csv(ifc_file: ifcopenshell.file, csv_filepath: Path) -> None:
        assert len(cost_schedules := ifc_file.by_type("IfcCostSchedule")) == 1
        all_cost_items = ifc_file.by_type("IfcCostItem")

        # Ensure everything was imported.
        rows = TestCsv2Ifc.get_items_rows_from_csv(csv_filepath)
        assert len(rows) == len(all_cost_items)

        # A rought test that some hierarchy was established.
        root_items = ifcopenshell.util.cost.get_root_cost_items(cost_schedules[0])
        assert len(all_cost_items) > len(root_items)
        all_nested_items = [i for item in root_items for i in ifcopenshell.util.cost.get_all_nested_cost_items(item)]
        assert len(all_nested_items + root_items) == len(all_cost_items)

    @pytest.mark.parametrize("csv_filepath", Path(__file__).parent.parent.glob("*.csv"))
    def test_import_sample_files(self, csv_filepath: Path):
        ifc_file = self.setup_ifc_file()
        ifcopenshell.api.root.create_entity(ifc_file, ifc_class="IfcProject")
        unit = ifcopenshell.api.unit.add_si_unit(ifc_file, unit_type="LENGTHUNIT", prefix="MILLI")
        ifcopenshell.api.unit.assign_unit(ifc_file, units=[unit])

        csv2ifc = ifc5d.csv2ifc.Csv2Ifc(str(csv_filepath), ifc_file)
        csv2ifc.execute()

        self.validate_ifc_file_against_csv(ifc_file, csv_filepath)

    def test_import_semicolon_delimited(self):
        # Regression test for #7048. Locales that use a comma as the decimal separator
        # use a semicolon as the field delimiter, which was previously unreadable.
        csv_filepath = Path(__file__).parent.parent / "sample_cost_schedule_house_FR.csv"
        rows = self.get_items_rows_from_csv(csv_filepath)

        with csv_filepath.open("r", encoding="utf-8") as csv_file:
            all_rows = list(csv.reader(csv_file))

        with tempfile.TemporaryDirectory("w") as temp_dir:
            semicolon_filepath = Path(temp_dir) / "semicolon.csv"
            with semicolon_filepath.open("w", newline="", encoding="utf-8") as csv_file:
                csv.writer(csv_file, delimiter=";").writerows(all_rows)

            # The comma reader must not silently treat the whole line as one column.
            assert len(self.get_items_rows_from_csv(semicolon_filepath, delimiter=";")) == len(rows)

            ifc_file = self.setup_ifc_file()
            csv2ifc = ifc5d.csv2ifc.Csv2Ifc(str(semicolon_filepath), ifc_file, delimiter=";")
            csv2ifc.execute()

            assert len(ifc_file.by_type("IfcCostSchedule")) == 1
            assert len(ifc_file.by_type("IfcCostItem")) == len(rows)

    def test_export_semicolon_delimited(self):
        ifc_file = self.setup_ifc_file()
        csv_filepath = Path(__file__).parent.parent / "sample_cost_schedule_house_FR.csv"
        ifc5d.csv2ifc.Csv2Ifc(str(csv_filepath), ifc_file).execute()

        with tempfile.TemporaryDirectory("w") as temp_csv_dir:
            writer = ifc5d.ifc5Dspreadsheet.Ifc5DCsvWriter(ifc_file, temp_csv_dir)
            writer.delimiter = ";"
            writer.write()

            exported = next(Path(temp_csv_dir).glob("*.csv"))
            n_rows = len(self.get_items_rows_from_csv(csv_filepath))
            assert len(self.get_items_rows_from_csv(exported, delimiter=";")) == n_rows

            # Round trip it back through the semicolon reader.
            new_ifc_file = self.setup_ifc_file()
            ifc5d.csv2ifc.Csv2Ifc(str(exported), new_ifc_file, delimiter=";").execute()
            assert len(new_ifc_file.by_type("IfcCostItem")) == n_rows

    def test_export_import_test(self):
        ifc_file = self.setup_ifc_file()

        csv_filepath = Path(__file__).parent.parent / "sample_cost_schedule_house_FR.csv"
        csv2ifc = ifc5d.csv2ifc.Csv2Ifc(str(csv_filepath), ifc_file)
        csv2ifc.execute()

        self.validate_ifc_file_against_csv(ifc_file, csv_filepath)
        n_rows = len(self.get_items_rows_from_csv(csv_filepath))

        # Export it and import it to different IFC file..
        with tempfile.TemporaryDirectory("w") as temp_csv_dir:
            writer = ifc5d.ifc5Dspreadsheet.Ifc5DCsvWriter(ifc_file, temp_csv_dir)
            writer.write()

            new_csv_filepath = next(Path(temp_csv_dir).glob("*.csv"))
            n_rows_ = len(self.get_items_rows_from_csv(new_csv_filepath))
            assert n_rows == n_rows_

            new_ifc_file = self.setup_ifc_file()
            csv2ifc = ifc5d.csv2ifc.Csv2Ifc(str(new_csv_filepath), new_ifc_file)
            csv2ifc.execute()

            self.validate_ifc_file_against_csv(new_ifc_file, new_csv_filepath)

    def test_export_xlsx_ods(self):
        ifc_file = self.setup_ifc_file()

        csv_filepath = Path(__file__).parent.parent / "sample_cost_schedule_house_FR.csv"
        csv2ifc = ifc5d.csv2ifc.Csv2Ifc(str(csv_filepath), ifc_file)
        csv2ifc.execute()

        self.validate_ifc_file_against_csv(ifc_file, csv_filepath)

        # Just to make sure it works.
        with tempfile.TemporaryDirectory("w") as temp_csv_dir:
            writer = ifc5d.ifc5Dspreadsheet.Ifc5DOdsWriter(ifc_file, temp_csv_dir)
            writer.write()
            writer = ifc5d.ifc5Dspreadsheet.Ifc5DXlsxWriter(ifc_file, temp_csv_dir)
            writer.write()
            assert len(list(Path(temp_csv_dir).glob("*.ods"))) == 1
            assert len(list(Path(temp_csv_dir).glob("*.xlsx"))) == 1

    def test_xlsx_columns_match_cost_panel(self):
        """ODS/XLSX are presentation formats: they must show exactly what the
        Bonsai cost panel shows (ID, Name, Quantity, Value, Total Cost), no
        internal bookkeeping columns, no Description/Unit, no per-category
        cost breakdown. See #6251."""
        import openpyxl

        ifc_file = self.setup_ifc_file()
        csv_filepath = Path(__file__).parent.parent / "sample_cost_schedule_house_FR.csv"
        ifc5d.csv2ifc.Csv2Ifc(str(csv_filepath), ifc_file).execute()

        with tempfile.TemporaryDirectory("w") as temp_dir:
            writer = ifc5d.ifc5Dspreadsheet.Ifc5DXlsxWriter(ifc_file, temp_dir)
            writer.write()
            workbook = openpyxl.load_workbook(next(Path(temp_dir).glob("*.xlsx")))
            worksheet = workbook.active

            headers = [cell.value for cell in next(worksheet.iter_rows())]
            assert headers == ["ID", "Name", "Quantity", "Value", "Total Cost"]

            # A leaf item (has quantity and value) gets Quantity * Value.
            leaf_row = next(row for row in worksheet.iter_rows(min_row=2) if row[0].value == "DB.1.1")
            assert leaf_row[4].value == "=C{}*D{}".format(leaf_row[0].row, leaf_row[0].row)

            # A parent/sum item gets the sum of its direct children's Total Cost.
            parent_row = next(row for row in worksheet.iter_rows(min_row=2) if row[0].value == "DB.1")
            assert parent_row[4].value.startswith("=SUM(")


class TestSerialiseCostQuantities:
    def test_quantity_name_with_special_characters_round_trips_as_json(self):
        ifc_file = ifcopenshell.file()
        name = 'Prospetto est "Np=256,667-23"'
        quantity = ifc_file.create_entity("IfcQuantityArea", Name=name, AreaValue=12.5)
        cost_item = ifc_file.create_entity("IfcCostItem", CostQuantities=[quantity])

        result = ifc5d.ifc5Dspreadsheet.IfcDataGetter.serialise_cost_quantities(ifc_file, cost_item)

        assert json.loads(result) == [[name, 12.5, ""]]

    def test_unset_name_does_not_crash(self):
        ifc_file = ifcopenshell.file()
        # Name left unset so quantity.Name resolves to None at access time.
        quantity = ifc_file.create_entity("IfcQuantityArea", AreaValue=3.0)
        cost_item = ifc_file.create_entity("IfcCostItem", CostQuantities=[quantity])

        result = ifc5d.ifc5Dspreadsheet.IfcDataGetter.serialise_cost_quantities(ifc_file, cost_item)

        assert json.loads(result) == [["", 3.0, ""]]

    def test_formula_is_included_when_present(self):
        ifc_file = ifcopenshell.file()
        quantity = ifc_file.create_entity("IfcQuantityArea", Name="Area", AreaValue=12.5, Formula="Length * Width")
        cost_item = ifc_file.create_entity("IfcCostItem", CostQuantities=[quantity])

        result = ifc5d.ifc5Dspreadsheet.IfcDataGetter.serialise_cost_quantities(ifc_file, cost_item)

        assert json.loads(result) == [["Area", 12.5, "Length * Width"]]

    def test_quantity_without_formula_attribute_does_not_crash(self):
        # IfcPhysicalComplexQuantity has no Formula attribute and is unsupported.
        ifc_file = ifcopenshell.file()
        quantity = ifc_file.create_entity("IfcPhysicalComplexQuantity", Name="Complex", Discrimination="layer")
        cost_item = ifc_file.create_entity("IfcCostItem", CostQuantities=[quantity])

        result = ifc5d.ifc5Dspreadsheet.IfcDataGetter.serialise_cost_quantities(ifc_file, cost_item)

        assert json.loads(result) == [["Complex ERROR: Only IfcPhysicalSimpleQuantity is supported", 0.0, ""]]
