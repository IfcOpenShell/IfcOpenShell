# IfcFM - IFC for facility management
# Copyright (C) 2023 Dion Moult <dion@thinkmoult.com>
#
# This file is part of IfcFM.
#
# IfcFM is free software: you can redistribute it and/or modify
# it under the terms of the GNU Lesser General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# IfcFM is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU Lesser General Public License for more details.
#
# You should have received a copy of the GNU Lesser General Public License
# along with IfcFM.  If not, see <http://www.gnu.org/licenses/>.

import os
import re
import csv
import importlib
import ifcopenshell.util.selector
from pathlib import Path
from collections import defaultdict
from typing import Literal, Union, Any, Callable

try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
except:
    pass  # No XLSX support

try:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.style import Style, TableCellProperties
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
except:
    pass  # No ODF support


try:
    import pandas as pd
except:
    pass  # No Pandas support

__version__ = version = "0.0.0"


ParserPreset = Literal["basic", "cobie24", "cobie24legacy"]
_parser_presets_configs = {}


def get_presets_configs() -> dict[ParserPreset, dict[str, Any]]:
    global _parser_presets_configs
    if not _parser_presets_configs:
        fm_dir = Path(__file__).parent
        for f in fm_dir.iterdir():
            if not f.suffix == ".py" or f.name.startswith("_"):
                continue
            preset = f.stem
            module = importlib.import_module(f"ifcfm.{preset}")
            config = getattr(module, "config")
            _parser_presets_configs[preset] = config
    return _parser_presets_configs


class Parser:
    config: dict[str, Any]
    categories: defaultdict[str, dict[str, Any]]
    get_custom_element_data: dict[
        str, Union[Callable[[ifcopenshell.file, ifcopenshell.entity_instance], dict[str, Any]], dict[str, Any]]
    ]
    duplicate_keys: list[tuple[dict[str, Any], dict[str, Any]]]

    def __init__(self, preset: Union[str, ParserPreset, dict[str, Any]] = "basic"):
        self.file = None
        self.preset = preset
        self.categories = defaultdict(dict)
        self.get_custom_element_data = {}
        self.duplicate_keys = []

        if isinstance(preset, str):
            presets = get_presets_configs()
            if preset not in presets:
                raise Exception(f"Invalid preset '{preset}'. Available presets: {','.join(presets.keys())}.")
            self.config = get_presets_configs()[preset]
        else:
            self.config = preset

    # TODO: name is unused?
    def parse(self, ifc_file: ifcopenshell.file, name=None):
        for category_name, category_config in self.config["categories"].items():
            for element in category_config["get_category_elements"](ifc_file):
                get_element_data = category_config["get_element_data"]

                if isinstance(get_element_data, dict):
                    data = {}
                    for key, query in get_element_data.items():
                        data[key] = ifcopenshell.util.selector.get_element_value(element, query)
                elif isinstance(get_element_data, Callable):
                    data = get_element_data(ifc_file, element) or {}

                get_custom_element_data = self.get_custom_element_data.get(category_name, lambda x, y: None)
                if isinstance(get_custom_element_data, dict):
                    custom_data = {}
                    for key, query in get_custom_element_data.items():
                        custom_data[key] = ifcopenshell.util.selector.get_element_value(element, query)
                elif isinstance(get_custom_element_data, Callable):
                    custom_data = get_custom_element_data(ifc_file, element) or {}

                data.update(custom_data)

                if data:
                    key = "-".join([str(data[k]) for k in category_config["keys"]])
                    # TODO: duplicate_keys are never used?
                    if key in self.categories[category_name]:
                        self.duplicate_keys.append((self.categories[category_name][key], data))
                    self.categories[category_name][key] = data

    def federate(self, paths: list[str]) -> None:
        for path in paths:
            spreadsheet = pd.ExcelFile(path)
            sheet_names = spreadsheet.sheet_names

            for category_name, category_config in self.config["categories"].items():
                if category_name not in sheet_names:
                    continue
                df = pd.read_excel(spreadsheet, sheet_name=category_name, keep_default_na=False)
                for _, row in df.iterrows():
                    key = "-".join([str(row[k]) for k in category_config["keys"]])
                    if key in self.categories[category_name]:
                        continue
                    self.categories[category_name][key] = row.to_dict()

    def exclude_categories(self, names: list[str]) -> None:
        for name in names:
            if name in self.config["categories"]:
                del self.config["categories"][name]

    def exclude_element_data(self, category: str, names: list[str]) -> None:
        headers = self.config["categories"][category]["headers"]
        self.config["categories"][category]["headers"] = [h for h in headers if h not in names]


class Writer:
    config: dict[str, Any]
    categories: dict[str, dict[str, Any]]

    def __init__(self, parser: Parser):
        self.parser = parser
        if isinstance(self.parser.preset, str):
            self.config = get_presets_configs()[self.parser.preset]
        elif isinstance(self.parser.preset, dict):
            self.config = self.parser.preset["config"]
        else:
            self.config = getattr(self.parser.preset, "config")

    def write(
        self,
        null: str = "N/A",
        empty: str = "-",
        bool_true: str = "YES",
        bool_false: str = "NO",
        list_separator: str = ", ",
    ) -> None:
        self.categories = {}
        null = self.config.get("null", null)
        empty = self.config.get("empty", empty)
        bool_true = self.config.get("bool_true", bool_true)
        bool_false = self.config.get("bool_false", bool_false)
        for category, config in self.config["categories"].items():
            data = self.parser.categories.get(category, None)
            headers = config["headers"]

            if not data:
                self.categories[category] = {"headers": headers, "rows": []}
                continue

            if not headers:
                headers = list(data[list(data.keys())[0]].keys())

            rows = []
            for row in data.values():
                processed_row = []
                for header in headers:
                    value = row[header]
                    if value is None:
                        value = null
                    elif value == "":
                        value = empty
                    elif value is True:
                        value = bool_true
                    elif value is False:
                        value = bool_false
                    elif isinstance(value, (list, tuple)):
                        value = list_separator.join([str(v) for v in value])
                    processed_row.append(value)
                rows.append(processed_row)

            sort = self.config.get("categories", {}).get(category, {}).get("sort", None)
            if sort:

                def natural_sort(value):
                    if isinstance(value, str):
                        convert = lambda text: int(text) if text.isdigit() else text.lower()
                        return [convert(c) for c in re.split("([0-9]+)", value)]
                    return [str(value)]

                # Sort least important keys first, then more important keys.
                # https://stackoverflow.com/questions/11476371/sort-by-multiple-keys-using-different-orderings
                for sort_data in reversed(sort):
                    i = headers.index(sort_data["name"])
                    reverse = sort_data["order"] == "DESC"
                    rows = sorted(rows, key=lambda x: natural_sort(x[i]), reverse=reverse)
            self.categories[category] = {"headers": headers, "rows": rows}

    def write_csv(self, output_folder: str, delimiter: str = ",") -> None:
        output_folder_ = Path(output_folder)
        output_folder_.mkdir(exist_ok=True)
        filename = None
        if len(self.categories.keys()) == 1 and "." in os.path.basename(output_folder_):
            filename = output_folder_
        for category, data in self.categories.items():
            category_filename = filename or output_folder_ / f"{category}.csv"
            with open(category_filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=delimiter)
                writer.writerow(data["headers"])
                for row in data["rows"]:
                    writer.writerow(row)

    def write_ods(self, output: str) -> None:
        doc = OpenDocumentSpreadsheet()

        for key, value in self.config.get("colours", {}).items():
            style = Style(name=key, family="table-cell")
            style.addElement(TableCellProperties(backgroundcolor="#" + value))
            doc.automaticstyles.addElement(style)

        for category, data in self.categories.items():
            colours = self.config.get("categories", {}).get(category, {}).get("colours", [])

            table = Table(name=category)
            tr = TableRow()
            for header in data["headers"]:
                tc = TableCell(valuetype="string", stylename="h")
                tc.addElement(P(text=header))
                tr.addElement(tc)
            table.addElement(tr)
            for row in data["rows"]:
                tr = TableRow()
                c = 0
                for col in row:
                    if c >= len(colours):
                        cell_format = "n"
                    else:
                        cell_format = colours[c]
                    tc = TableCell(valuetype="string", stylename=cell_format)
                    tc.addElement(P(text=str(col)))
                    tr.addElement(tc)
                    c += 1
                table.addElement(tr)
            doc.spreadsheet.addElement(table)

        if len(output) > 4 and output[-4:].lower() == ".ods":
            output = output[0:-4]

        doc.save(output, True)

    def write_xlsx(self, output: str) -> None:
        workbook = Workbook()

        cell_formats = {}
        for key, value in self.config.get("colours", {}).items():
            fill = PatternFill(start_color=value, end_color=value, fill_type="solid")
            cell_formats[key] = fill

        for category, data in self.categories.items():
            colours = self.config.get("categories", {}).get(category, {}).get("colours", [])

            if category in workbook.sheetnames:
                worksheet = workbook[category]
            else:
                worksheet = workbook.create_sheet(category)

            r = 1  # Openpyxl uses 1-based indexing
            c = 1
            for header in data["headers"]:
                cell = worksheet.cell(row=r, column=c, value=header)
                cell.fill = cell_formats["h"]
                c += 1

            r += 1
            for row in data["rows"]:
                c = 1
                for col in row:
                    if c > len(colours):  # Adjusted the comparison
                        cell_format = "n"
                    else:
                        cell_format = colours[c - 1]  # Adjusted the indexing
                    cell = worksheet.cell(row=r, column=c, value=str(col))
                    cell.fill = cell_formats[cell_format]
                    c += 1
                r += 1

        if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
            workbook.remove(workbook["Sheet"])

        workbook.save(output)

    def write_pd(self) -> dict[str, pd.DataFrame]:
        results = {}
        for category, data in self.categories.items():
            results[category] = pd.DataFrame(data["rows"], columns=data["headers"])
        return results
