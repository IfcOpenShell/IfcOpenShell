# Bonsai - OpenBIM Blender Add-on
# Copyright (C) 2020, 2021 Dion Moult <dion@thinkmoult.com>
#
# This file is part of Bonsai.
#
# Bonsai is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# Bonsai is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Bonsai.  If not, see <http://www.gnu.org/licenses/>.

import os
import re
import bpy
import string
import svgwrite
import openpyxl

from bonsai.bim.module.drawing.svgwriter import SvgWriter
from odf.opendocument import load as load_ods
from odf.table import Table, TableRow, TableColumn, TableCell
from odf.text import P
from odf.style import Style
from textwrap import wrap
from pathlib import Path
from bonsai.bim.ifc import IfcStore

DEBUG = False


def col2num(col):
    """convert letter column index to number:
    `"A" -> 1`, `"AA" -> 27``
    """
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord("A")) + 1
    return num


def a1_to_rc(cell):
    """convert cell index from A1 format to RC: `"A1" -> (0,0)`"""
    column_letter = cell.strip(string.digits)
    col_number = col2num(column_letter) - 1
    row_number = int(cell[len(column_letter) :]) - 1
    return row_number, col_number


class Scheduler:
    def schedule(self, infile, outfile):
        self.svg = svgwrite.Drawing(
            outfile,
            debug=False,
            id="root",
        )
        self.padding = 1
        self.margin = 1

        self.parse_css(infile)
        if infile.endswith("ods"):
            self.schedule_ods(infile, outfile)
        elif infile.endswith("xlsx"):
            self.schedule_xlsx(infile, outfile)

    def parse_css(self, infile):
        stylesheet_path = os.path.splitext(infile)[0] + ".css"
        if not os.path.exists(stylesheet_path):
            stylesheet_rel_path = getattr(bpy.context.scene.DocProperties, "schedules_stylesheet_path")
            ifc_file_path = os.path.dirname(IfcStore.path)
            stylesheet_path = ifc_file_path + "\\" + stylesheet_rel_path
            if not os.path.exists(stylesheet_path):
                stylesheet_path = os.path.join(bpy.context.scene.BIMProperties.data_dir, "assets", "schedule.css")
        with open(stylesheet_path, "r") as stylesheet:
            css = stylesheet.read()

        matches = re.search(r"--font-size-pt:\s*([0-9.]+);", css)
        self.font_size_pt = float(matches.groups()[0]) if matches else 12  # Default to 12pt
        matches = re.search(r"--font-size-px:\s*([0-9.]+);", css)
        self.font_size_px = float(matches.groups()[0]) if matches else 4.13  # Magic number 4.13px ~= 12pt
        matches = re.search(r"--font-width:\s*([0-9.]+);", css)
        self.font_width = float(matches.groups()[0]) if matches else 0.45  # A magic number for OpenGost

        self.svg.defs.add(self.svg.style(css))

    def schedule_xlsx(self, infile, outfile):
        workbook = openpyxl.open(infile, data_only=True)
        sheet = workbook.active

        column_dimensions = {k: v.width for k, v in sheet.column_dimensions.items()}
        row_dimensions = {k: v.height for k, v in sheet.row_dimensions.items()}
        default_width = 8.43  # 8.43 characters wide
        default_height = 15  # 15pt seems to be the default height

        merged_cells = []
        for cell_range in sheet.merged_cells.ranges:
            pass

        y = self.margin
        rows = list(sheet.iter_rows())
        total_rows = len(rows)
        for i, row in enumerate(rows):
            # The last row may contain only null values
            if i == (total_rows - 1) and not [c for c in row if c.value is not None]:
                continue

            x = self.margin
            for cell in row:
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    column_letter = openpyxl.utils.get_column_letter(cell.column)
                    width = column_dimensions.get(column_letter, default_width)
                else:
                    width = column_dimensions.get(cell.column_letter, default_width)
                width = self.convert_character_width_to_mm(width)
                height = row_dimensions.get(cell.row, default_height)
                height = self.convert_to_mm(f"{height}pt")

                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    x += width
                    continue

                unmerged_width = width
                unmerged_height = height

                for cell_range in sheet.merged_cells.ranges:
                    if cell.coordinate not in cell_range:
                        continue
                    merged_width = 0
                    merged_height = 0
                    for merged_column in cell_range.cols:
                        merged_cell = sheet.cell(*merged_column[0])
                        column_letter = openpyxl.utils.get_column_letter(merged_cell.column)
                        width = column_dimensions.get(cell.column_letter, default_width)
                        width = self.convert_character_width_to_mm(width)
                        merged_width += width
                    for merged_row in cell_range.rows:
                        merged_cell = sheet.cell(*merged_row[0])
                        height = row_dimensions.get(merged_cell.row, default_height)
                        height = self.convert_to_mm(f"{height}pt")
                        merged_height += height
                    width = merged_width
                    height = merged_height
                    break

                if cell.fill.patternType:
                    background_color = "#" + str(cell.fill.bgColor.rgb).lower()[2:]
                else:
                    background_color = "#ffffff"

                self.svg.add(
                    self.svg.rect(
                        insert=(x, y),
                        size=(width, height),
                        style=f"fill: {background_color};",
                        class_="border",
                    )
                )

                if not cell.value:
                    x += unmerged_width
                    continue

                font_size = cell.font.size or self.font_size_pt  # 12pt default
                font_size = font_size / self.font_size_pt * self.font_size_px  # Magic?

                text_position = [0.0, 0.0]
                if cell.alignment.horizontal == "left":
                    text_position[0] = x + self.padding
                    horizontal_align = "left"
                elif cell.alignment.horizontal == "general" and isinstance(cell.value, str):
                    text_position[0] = x + self.padding
                    horizontal_align = "left"
                elif cell.alignment.horizontal in ("center", "justify", "centerContinuous", "distributed"):
                    text_position[0] = x + width / 2
                    horizontal_align = "middle"
                elif cell.alignment.horizontal == "right":
                    text_position[0] = x + width - self.padding
                    horizontal_align = "right"
                elif cell.alignment.horizontal == "general" and not isinstance(cell.value, str):
                    text_position[0] = x + width - self.padding
                    horizontal_align = "right"
                else:
                    text_position[0] = x + self.padding
                    horizontal_align = "left"

                if cell.alignment.vertical == "top":
                    text_position[1] = y + self.padding
                    vertical_align = "top"
                elif cell.alignment.vertical in ("center", "justify", "distributed"):
                    text_position[1] = y + height / 2
                    vertical_align = "middle"
                elif cell.alignment.vertical == "bottom":
                    text_position[1] = y + height - self.padding
                    vertical_align = "bottom"
                else:
                    text_position[1] = y + height - self.padding
                    vertical_align = "bottom"

                if vertical_align == "middle" and horizontal_align == "middle":
                    box_alignment = "center"
                else:
                    box_alignment = f"{vertical_align}-{horizontal_align}"

                if cell.font.color:
                    text_color = "#" + str(cell.font.color.rgb).lower()[2:]
                else:
                    text_color = "#000000"

                self.add_text(
                    [cell.value],
                    *text_position,
                    font_size=font_size,
                    box_alignment=box_alignment,
                    wrap_text=cell.alignment.wrapText or False,
                    cell_width=width,
                    bold=cell.font.b,
                    italic=cell.font.i,
                    text_color=text_color,
                )

                x += unmerged_width
            y += unmerged_height

        total_width = x + self.margin
        total_height = y + self.margin
        self.svg["width"] = "{}mm".format(total_width)
        self.svg["height"] = "{}mm".format(total_height)
        self.svg["viewBox"] = "0 0 {} {}".format(total_width, total_height)
        self.svg.save(pretty=True)

    def schedule_ods(self, infile, outfile):
        doc = load_ods(infile)

        # useful for debugging ods
        if DEBUG:
            import xml.dom.minidom

            path = Path(infile)
            dom = xml.dom.minidom.parseString(doc.xml())
            pretty_xml = dom.toprettyxml()

            with open(path.with_suffix(".xml"), "w") as fo:
                fo.write(pretty_xml)

        styles = {}
        for cell_style in doc.getElementsByType(Style):
            name = cell_style.getAttribute("name")
            styles[name] = {}

            # NOTE: there are also styles that inherit from parent styles that we do not process atm
            if not cell_style.firstChild:
                continue

            if cell_style.firstChild.tagName in ["style:table-column-properties", "style:table-row-properties"]:
                style_children = [cell_style.firstChild]
            else:
                # for style:table-cell-properties we need to collect also text and paragraph properties
                style_children = cell_style.childNodes

            for child in style_children:
                child_params = {key[1]: value for key, value in child.attributes.items()}
                styles[name].update(child_params)

        table = doc.getElementsByType(Table)[0]

        # related styles stored as a list of tuples:
        # [(child, parent), ...]
        related_styles = []

        # collect columns width
        column_widths = []
        column_styles = []
        for col in table.getElementsByType(TableColumn):
            style_name = col.getAttribute("stylename")
            col_repeat = col.getAttribute("numbercolumnsrepeated")
            col_repeat = int(col_repeat) if col_repeat else 1
            for i in range(col_repeat):
                if not style_name or "column-width" not in styles[style_name]:
                    column_width = 50
                else:
                    column_width = self.convert_to_mm(styles[style_name]["column-width"])
                column_styles.append(style_name)
                column_widths.append(column_width)
            cell_style = col.getAttribute("defaultcellstylename")
            if cell_style:
                related_styles.append((style_name, cell_style))

        # sometimes there are no column styles (e.g. in ODS from IfcCSV)
        # and we just use some constant number for column widths
        if not column_widths:
            row_columns = []
            for tr in table.getElementsByType(TableRow):
                row_cols = 0
                for td in tr.getElementsByType(TableCell):
                    column_span = td.getAttribute("numbercolumnsspanned")
                    column_span = int(column_span) if column_span else 1

                    col_repeat = td.getAttribute("numbercolumnsrepeated")
                    col_repeat = int(col_repeat) if col_repeat else 1
                    row_cols += column_span * col_repeat
                row_columns.append(row_cols)

            n_columns = max(row_columns)
            column_widths = [25] * n_columns  # some constant width value 👀
            column_styles = [None] * n_columns

        # collect rows height
        row_heights = []
        # TODO: never used yet because unsure about priority for row styles
        # over column styles or vice versa
        row_styles = []

        for col in table.getElementsByType(TableRow):
            style_name = col.getAttribute("stylename")
            row_repeat = col.getAttribute("numberrowsrepeated")
            row_repeat = int(row_repeat) if row_repeat else 1
            for i in range(row_repeat):
                if not style_name or "row-height" not in styles[style_name]:
                    row_height = 6
                else:
                    row_height = self.convert_to_mm(styles[style_name]["row-height"])
                row_styles.append(style_name)
                row_heights.append(row_height)
            cell_style = col.getAttribute("defaultcellstylename")
            if cell_style:
                related_styles.append((style_name, cell_style))

        while len(related_styles) > 0:
            # unzip related styles to children and parents
            children, parents = zip(*related_styles)
            independent_styles = set(parents) - set(children)
            for relation in related_styles[:]:
                child, parent = relation
                if parent in independent_styles:
                    child_style = styles[child]
                    styles[child] = styles[parent] | child_style
                    related_styles.remove(relation)

        # TODO: multiple print ranges? 😔
        print_range = table.getAttribute("printranges")
        if print_range:
            min_rc, max_rc = [a1_to_rc(cell.rsplit(".", 1)[1]) for cell in print_range.split(":")]
        else:
            # fallback if print range is not defined
            n_rows = len(row_heights)
            n_cols = len(column_widths)
            n_cells = len(row_heights) * len(column_widths)
            cells_limit = 10000
            if n_cells >= cells_limit:
                raise Exception(
                    f"You were about to build a very big table with number of cells more than {cells_limit}.\n"
                    f"In fact it is {n_rows} rows x {n_cols} cols = {n_cells} cells \n"
                    "and the operation was stopped to prevent system freeze.\n"
                    "Please define print range in .ods file to proceede\n"
                    "(needed to make sure printed table will have reasonable size)."
                )

            min_rc, max_rc = (0, 0), (1048576, 16384)
        min_row, min_col = min_rc
        max_row, max_col = max_rc

        # draw table
        y = self.margin
        tri = 0
        stop_iterating_over_rows = False
        # TODO: row spans support?
        for tr in table.getElementsByType(TableRow):
            if stop_iterating_over_rows:
                break

            row_repeat = tr.getAttribute("numberrowsrepeated")
            row_repeat = int(row_repeat) if row_repeat else 1

            for i_row_repeat in range(row_repeat):
                if tri < min_row:
                    tri += 1
                    continue
                elif tri > max_row:
                    stop_iterating_over_rows = True
                    break

                x = self.margin
                height = row_heights[tri]
                tdi = 0
                stop_iterating_over_columns = False

                for td in tr.getElementsByType(TableCell):
                    if stop_iterating_over_columns:
                        break

                    column_span = td.getAttribute("numbercolumnsspanned")
                    column_span = int(column_span) if column_span else 1

                    col_repeat = td.getAttribute("numbercolumnsrepeated")
                    col_repeat = int(col_repeat) if col_repeat else 1

                    # figuring text alignment
                    cell_style = self.get_style(td.getAttribute("stylename"), styles)

                    # drawing cells and text
                    for i_col_repeat in range(col_repeat):
                        start_tdi = tdi
                        end_tdi = tdi + int(column_span) - 1
                        # if the entire span is beyond print range => continue
                        # if only part then keeping that part
                        if start_tdi < min_col:
                            if end_tdi < min_col:
                                tdi += column_span
                                continue
                            else:
                                start_tdi = min_col

                        # stop if start column is beyond print range
                        if start_tdi > max_col:
                            stop_iterating_over_columns = True
                            break

                        # making sure last column won't go beyond the print range
                        if end_tdi > max_col:
                            end_tdi = max_col

                        width = sum(column_widths[start_tdi : end_tdi + 1])
                        col_style = self.get_style(column_styles[tdi], styles)
                        final_cell_style = cell_style or col_style
                        background_color = final_cell_style.get("background-color", "#ffffff")

                        self.svg.add(
                            self.svg.rect(
                                insert=(x, y),
                                size=(width, height),
                                style=f"fill: {background_color};",
                                class_="border",
                            )
                        )

                        p_tags = td.getElementsByType(P)
                        text_color = final_cell_style.get("color", None)
                        box_alignment = self.get_box_alignment(final_cell_style)
                        wrap_text = final_cell_style.get("wrap-option", None) == "wrap"
                        bold_text = final_cell_style.get("font-weight", None) == "bold"
                        italic_text = final_cell_style.get("font-style", None) == "italic"
                        # NOTE: very naive since we're scaling text proportionally
                        font_size = (
                            float(final_cell_style.get("font-size", f"{self.font_size_pt}pt")[:-2])
                            / self.font_size_pt
                            * self.font_size_px
                        )

                        if p_tags:
                            # figuring text position based on alignment
                            text_position = [0.0, 0.0]
                            if box_alignment.endswith("left"):
                                text_position[0] = x + self.padding
                            elif box_alignment.endswith("middle") or box_alignment == "center":
                                text_position[0] = x + width / 2
                            elif box_alignment.endswith("right"):
                                text_position[0] = x + width - self.padding

                            if box_alignment.startswith("top"):
                                text_position[1] = y + self.padding
                            elif box_alignment.startswith("middle") or box_alignment == "center":
                                text_position[1] = y + height / 2
                            elif box_alignment.startswith("bottom"):
                                text_position[1] = y + height - self.padding

                            self.add_text(
                                p_tags,
                                *text_position,
                                font_size=font_size,
                                box_alignment=box_alignment,
                                wrap_text=wrap_text,
                                cell_width=width,
                                bold=bold_text,
                                italic=italic_text,
                                text_color=text_color,
                            )
                        x += width
                        tdi += column_span

                tri += 1
                y += height

        total_width = x + self.margin
        total_height = y + self.margin
        self.svg["width"] = "{}mm".format(total_width)
        self.svg["height"] = "{}mm".format(total_height)
        self.svg["viewBox"] = "0 0 {} {}".format(total_width, total_height)
        self.svg.save(pretty=True)

    def get_style(self, style_name, styles):
        style = styles[style_name] if style_name else {}
        return style

    def get_box_alignment(self, style):
        if style and "vertical-align" in style and style["vertical-align"] != "automatic":
            vertical_align = style["vertical-align"]
        else:
            vertical_align = "bottom"

        alignment_translation = {
            "center": "middle",
            "end": "right",
            "start": "left",
        }

        if style and "text-align" in style and style["text-align"] != "automatic":
            horizontal_align = style["text-align"]
            horizontal_align = alignment_translation.get(horizontal_align, horizontal_align)
        else:
            horizontal_align = "left"

        if vertical_align == "middle" and horizontal_align == "middle":
            box_alignment = "center"
        else:
            box_alignment = f"{vertical_align}-{horizontal_align}"

        return box_alignment

    def add_text(
        self,
        p_tags,
        x,
        y,
        font_size,
        box_alignment="bottom-left",
        wrap_text=False,
        cell_width=100,
        bold=False,
        italic=False,
        text_color=None,
    ):
        """
        Adds text to svg.

        Args:
            p_tags: list of cell's P tags from odt file
            box_alignment: alignment of text in box
            wrap_text: if True, text will be wrapped to fit in cell
            cell_width: width of cell, used for wrapping text
        """
        text_lines = [str(p) for p in p_tags]
        box_alignment_params = SvgWriter.get_box_alignment_parameters(box_alignment)
        text_params = {"font-size": font_size}
        if bold:
            text_params["font-weight"] = "bold"
        if italic:
            text_params["font-style"] = "italic"
        if text_color:
            text_params["fill"] = text_color

        text_params["class_"] = "schedule"

        if len(text_lines) == 1 and not wrap_text:
            text_params.update(box_alignment_params)
            text_tag = self.svg.text(text_lines[0], insert=(x, y), **(text_params))
            if self.is_url(text_lines[0]):
                anchor = self.svg.a(text_lines[0].strip())
                anchor.add(text_tag)
                text_tag = anchor
            self.svg.add(text_tag)
            return

        text_tag = self.svg.text("", **(text_params | {"font-size": "0"} | box_alignment_params))

        # TODO: Should be done without using magic number for self.font_width
        if wrap_text:
            wrapped_lines = []
            for line in text_lines:
                wrapped_line = wrap(
                    line, width=int(cell_width // (font_size * self.font_width)), break_long_words=False
                )
                wrapped_lines.extend(wrapped_line)
        else:
            wrapped_lines = text_lines

        if box_alignment.startswith("top"):
            dy_dir = 1
            line_number = 0
        elif box_alignment.startswith("bot"):
            dy_dir = -1
            line_number = 0
        else:  # middle row
            # the idea is that the middle row should always stay in the center
            # e.g. dy offset for 3 lines is: 1, 0, -1
            #                for 2 lines:    0.5, -0.5
            dy_dir = -1
            line_number = (len(wrapped_lines) - 1) / 2

        for text_line in wrapped_lines[::dy_dir]:
            # position has to be inserted at tspan to avoid x offset between tspans
            tspan = self.svg.tspan(text_line, insert=(x, y), **text_params)
            # doing it here and not in tspan constructor because constructor adds unnecessary spaces
            tspan.update({"dy": f"{line_number}em"})
            if self.is_url(text_line):
                anchor = self.svg.a(text_line.strip())
                anchor.add(tspan)
                tspan = anchor
            text_tag.add(tspan)
            line_number += dy_dir

        self.svg.add(text_tag)

    def is_url(self, value):
        return value.strip().startswith("http") and " " not in value.strip() and "://" in value  # Wrong but fast

    def convert_character_width_to_mm(self, value):
        # Excel measures widths in terms of "number of characters of the
        # default font", which is most commonly 11pt Calibri. The width of the
        # "0" character in Calibri is approximately 7px at 100% zoom.
        px_per_character = 7
        px_to_pt = 0.75
        return self.convert_to_mm(f"{value * px_per_character * px_to_pt}pt")

    def convert_to_mm(self, value):
        # XSL is what defines the units of measurements in ODF
        # https://docs.oasis-open.org/office/v1.2/os/OpenDocument-v1.2-os-part1.html#datatype-positiveLength
        # https://www.w3.org/TR/2001/REC-xsl-20011015/slice5.html#section-N8185-Definitions-of-Units-of-Measure
        if "cm" in value:
            return float(value[0:-2]) * 10
        elif "mm" in value:
            return float(value[0:-2])
        elif "in" in value:
            return float(value[0:-2]) * 25.4
        elif "pt" in value:
            return float(value[0:-2]) * (1 / 72) * 25.4
        elif "pc" in value:
            return float(value[0:-2]) * 12 * (1 / 72) * 25.4
        elif "px" in value:
            # implementors may instead simply pick a fixed conversion factor,
            # treating 'px' as an absolute unit of measurement (such as 1/92" or
            # 1/72"). <-- We're picking 1/96 to match SVG. Let me know if it
            # breaks anything.
            return float(value[0:-2]) * (1 / 96) * 2.54 * 10
        elif "em" in value:
            # This is a funny one. Since the scheduler at the moment enforces a
            # font size of 2.5mm (vertically, FWIW), I'm writing this. Hopefully
            # this code doesn't hurt anybody.
            return float(value[0:-2]) * (1 / 96) * 2.54 * 10
